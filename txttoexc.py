import pandas as pd
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def extraire_infos_clients(contenu):
    """
    Extrait les informations des clients à partir du contenu du fichier texte
    """
    clients = []
    
    # Séparation par client (blocs de texte)
    blocs_clients = contenu.strip().split('\n\n')
    
    for bloc in blocs_clients:
        client_actuel = {}
        lignes = bloc.strip().split('\n')
        
        for ligne in lignes:
            # Utilisation d'expressions régulières pour extraire les paires clé-valeur
            match = re.match(r'([^:]+)\s*:\s*(.*)', ligne)
            if match:
                cle = match.group(1).strip()
                valeur = match.group(2).strip()
                
                if cle == "Nom":
                    client_actuel["Nom_Entreprise"] = valeur
                elif cle == "Secteur":
                    client_actuel["Secteur"] = valeur
                elif cle == "Localisation":
                    # Séparation de la ville et du pays
                    client_actuel["Localisation"] = valeur
                    if "," in valeur:
                        ville, pays = valeur.split(",", 1)
                        client_actuel["Ville"] = ville.strip()
                        client_actuel["Pays"] = pays.strip()
                    else:
                        client_actuel["Ville"] = ""
                        client_actuel["Pays"] = valeur
                elif cle == "Description":
                    client_actuel["Description"] = valeur
                elif cle == "Contact":
                    client_actuel["Contact"] = valeur
                    # Extraction du nom et du poste
                    if "(" in valeur and ")" in valeur:
                        nom_complet = valeur.split("(")[0].strip()
                        poste = valeur.split("(")[1].split(")")[0].strip()
                        client_actuel["Nom_Contact"] = nom_complet
                        client_actuel["Poste_Contact"] = poste
                    else:
                        client_actuel["Nom_Contact"] = valeur
                        client_actuel["Poste_Contact"] = ""
        
        if client_actuel:
            clients.append(client_actuel)
    
    return clients

def formater_excel(workbook, sheet_name):
    """
    Applique un formatage professionnel au fichier Excel
    """
    worksheet = workbook[sheet_name]
    
    # Définir les styles
    header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    
    # Appliquer le style à l'en-tête
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Appliquer des bordures et l'alignement à toutes les cellules
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # Ajuster la hauteur de la première ligne
    worksheet.row_dimensions[1].height = 25
    
    # Coloration alternée des lignes
    for row_idx in range(2, worksheet.max_row + 1):
        if row_idx % 2 == 0:  # Lignes paires
            light_fill = PatternFill(start_color="E6EFF9", end_color="E6EFF9", fill_type="solid")
            for cell in worksheet[row_idx]:
                cell.fill = light_fill
    
    # Figer les volets sur la première ligne
    worksheet.freeze_panes = "A2"
    
    return workbook

def convertir_en_excel(clients, fichier_sortie="clients.xlsx"):
    """
    Convertit les données des clients en fichier Excel avec formatage
    """
    # Création du DataFrame
    df = pd.DataFrame(clients)
    
    # Réorganiser les colonnes dans un ordre logique
    colonnes_ordre = [
        "Nom_Entreprise", "Secteur", "Pays", "Ville", 
        "Nom_Contact", "Poste_Contact", "Description"
    ]
    
    # S'assurer que toutes les colonnes existent
    for col in colonnes_ordre:
        if col not in df.columns:
            df[col] = ""
    
    # Réordonner les colonnes
    df = df[colonnes_ordre]
    
    # Renommer les colonnes pour l'affichage
    colonnes_renommees = {
        "Nom_Entreprise": "Entreprise",
        "Secteur": "Secteur",
        "Pays": "Pays",
        "Ville": "Ville",
        "Nom_Contact": "Contact",
        "Poste_Contact": "Poste",
        "Description": "Description"
    }
    df = df.rename(columns=colonnes_renommees)
    
    # Créer un objet ExcelWriter
    with pd.ExcelWriter(fichier_sortie, engine='openpyxl') as writer:
        # Convertir le DataFrame en une feuille Excel
        df.to_excel(writer, sheet_name='Clients IBM', index=False)
        
        # Récupérer le classeur
        workbook = writer.book
        
        # Formater le classeur
        workbook = formater_excel(workbook, 'Clients IBM')
        
        # Accéder à la feuille de calcul
        worksheet = workbook['Clients IBM']
        
        # Ajuster la largeur des colonnes
        for i, col in enumerate(df.columns):
            # Trouver la longueur maximale dans la colonne
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(col)
            )
            # Ajouter un peu d'espace supplémentaire
            adjusted_width = min(max_length + 3, 50)  # Limiter à 50 caractères max
            # Définir la largeur de colonne
            col_letter = get_column_letter(i + 1)
            worksheet.column_dimensions[col_letter].width = adjusted_width
        
        # Définir une largeur plus grande pour la colonne Description
        worksheet.column_dimensions[get_column_letter(7)].width = 60
    
    print(f"Fichier Excel créé avec succès: {fichier_sortie}")
    return fichier_sortie

def main():
    try:
        # Chemin du fichier d'entrée
        fichier_entree = "clients.txt"
        fichier_sortie = "clients_ibm.xlsx"
        
        # Lire le contenu du fichier
        with open(fichier_entree, 'r', encoding='utf-8') as f:
            contenu = f.read()
        
        # Extraire les informations des clients
        clients = extraire_infos_clients(contenu)
        
        # Convertir en Excel
        fichier_excel = convertir_en_excel(clients, fichier_sortie)
        
        print(f"Conversion terminée avec succès. {len(clients)} clients exportés dans {fichier_excel}.")
    
    except Exception as e:
        print(f"Erreur lors de la conversion: {e}")

if __name__ == "__main__":
    main()